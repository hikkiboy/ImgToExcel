from openpyxl import workbook, load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from PIL import Image

def rgb2hex(r, g, b):
    return '{:02x}{:02x}{:02x}'.format(r, g, b)

img = Image.open('Aphelios.jpg')
color = []
count = 0
visited = []
wb = load_workbook('Book1.xlsx')
ws = wb.active

if img.mode in ('RGB', 'LA') or (img.mode == 'P' and 'transparency' in img.info):   
    pixels = list(img.convert('RGBA').getdata())
    for r, g, b, a in pixels: 
                    color.append('00' + rgb2hex(r,g,b)) 
                    
                    
for row in range(1, 119):
    for col in range(1,201):           
        for x in color:
                    print(count)
                    cell = ws.cell(row=row ,column=col)
                    visited.append(cell)
                    count += 1
                    try:
                        ws.cell(row=row, column= col).fill = PatternFill(start_color=color[count],
                                                            end_color=color[count],
                                                            fill_type='solid')
                        print(ws.cell(row=row, column=col))
                        break
                    except:
                        break
                        
print('Finished')
wb.save('Book1.xlsx')


















